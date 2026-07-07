import io

from fastapi import HTTPException, UploadFile
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"}


def _extract_pdf(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages).strip()


def _extract_docx(content: bytes) -> str:
    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs).strip()


def _extract_excel(content: bytes) -> str:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                rows.append(" | ".join(cells))
    wb.close()
    return "\n".join(rows).strip()


async def extract_text(file: UploadFile) -> str:
    """Yüklenen dosyadan metin çıkarır. PDF, DOCX ve Excel desteklenir."""
    filename = file.filename or ""
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Desteklenmeyen dosya formatı: '{ext}'. "
                   f"Desteklenen formatlar: {', '.join(sorted(SUPPORTED_EXTENSIONS))}",
        )

    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Dosya boş")

    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".xlsx": _extract_excel,
        ".xls": _extract_excel,
    }

    text = extractors[ext](content)

    if not text:
        raise HTTPException(status_code=400, detail="Dosyadan metin çıkarılamadı")

    return text
