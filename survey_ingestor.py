from __future__ import annotations

import csv
import io
import math
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from fastapi import HTTPException, UploadFile


SUPPORTED_SURVEY_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_ROWS = 10_000
MAX_COLUMNS = 300

DEMOGRAPHIC_KEYWORDS = {
    "no",
    "tarih",
    "ikamet",
    "mahalle",
    "ilce",
    "ilçe",
    "cinsiyet",
    "yas",
    "yaş",
    "ogrenim",
    "öğrenim",
    "meslek",
    "engel",
    "bolum",
    "bölüm",
    "kutuphane",
    "kütüphane",
    "salon",
    "hizmet binasi",
    "hizmet binası",
}


@dataclass
class ParsedColumn:
    name: str
    original_name: str
    dtype: str
    semantic_type: str
    missing_count: int
    missing_pct: float
    unique_count: int
    sample_values: list[Any] = field(default_factory=list)
    code_map: dict[str, str] = field(default_factory=dict)
    order_index: int = 0


@dataclass
class ParsedQuestion:
    column_name: str
    question_no: str | None
    question_text: str
    question_type: str
    scale_type: str | None
    options: dict[str, str] = field(default_factory=dict)
    is_likert: bool = False
    is_demographic: bool = False
    is_open_text: bool = False
    order_index: int = 0


@dataclass
class SurveyReportData:
    summary: dict[str, Any]
    metrics: dict[str, Any]
    quality_issues: list[dict[str, Any]]
    report_text: str


@dataclass
class ParsedSurvey:
    filename: str
    file_type: str
    title: str
    source_sheet: str
    header_row: int
    data_start_row: int
    row_count: int
    column_count: int
    columns: list[ParsedColumn]
    questions: list[ParsedQuestion]
    rows: list[dict[str, Any]]
    report: SurveyReportData
    department: str | None = None
    period: str | None = None
    quarter: str | None = None
    year: int | None = None


async def parse_survey_upload(file: UploadFile, content: bytes | None = None) -> ParsedSurvey:
    filename = file.filename or "unknown"
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_SURVEY_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Desteklenmeyen dosya formatı: {ext}. Desteklenen formatlar: .csv, .xlsx, .xls",
        )

    if content is None:
        content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Dosya boş")
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Dosya boyutu izin verilen sınırı aşıyor")

    if ext == ".csv":
        sheet_name, raw_rows = _read_csv_rows(content)
    elif ext == ".xlsx":
        sheet_name, raw_rows = _read_xlsx_rows(content)
    else:
        sheet_name, raw_rows = _read_xls_rows(content)

    if not raw_rows:
        raise HTTPException(status_code=400, detail="Dosyada işlenebilir veri bulunamadı")

    title = _detect_title(raw_rows) or Path(filename).stem
    header_index = _detect_header_row(raw_rows)
    if header_index is None:
        raise HTTPException(status_code=422, detail="Anket başlık satırı otomatik tespit edilemedi")

    code_row = raw_rows[header_index - 1] if header_index > 0 else []
    headers = _dedupe_headers(raw_rows[header_index])
    if len(headers) > MAX_COLUMNS:
        raise HTTPException(status_code=413, detail="Kolon sayısı izin verilen sınırı aşıyor")

    data_rows = _build_data_rows(raw_rows[header_index + 1 :], headers)
    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=413, detail="Satır sayısı izin verilen sınırı aşıyor")
    if not data_rows:
        raise HTTPException(status_code=400, detail="Anket cevap satırı bulunamadı")

    columns = _build_columns(headers, code_row, data_rows)
    questions = _build_questions(columns)
    report = _build_report(title, columns, questions, data_rows)

    return ParsedSurvey(
        filename=filename,
        file_type=ext.lstrip("."),
        title=title,
        source_sheet=sheet_name,
        header_row=header_index + 1,
        data_start_row=header_index + 2,
        row_count=len(data_rows),
        column_count=len(headers),
        columns=columns,
        questions=questions,
        rows=data_rows,
        report=report,
        department=_detect_department(title),
        period=_detect_period(filename),
        quarter=_detect_quarter(filename),
        year=_detect_year(filename),
    )


def _read_csv_rows(content: bytes) -> tuple[str, list[list[Any]]]:
    text = _decode_csv_text(content)
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = "\t" if "\t" in first_line else ";" if ";" in first_line else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [[_normalize_cell(cell) for cell in row] for row in reader]
    return "CSV", _trim_empty_edges(rows)


def _read_xlsx_rows(content: bytes) -> tuple[str, list[list[Any]]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel dosyası okunamadı: {exc}") from exc

    try:
        ws = _select_worksheet(wb)
        rows = [
            [_normalize_cell(value) for value in row]
            for row in ws.iter_rows(values_only=True)
        ]
        return ws.title, _trim_empty_edges(rows)
    finally:
        wb.close()


def _read_xls_rows(content: bytes) -> tuple[str, list[list[Any]]]:
    try:
        xls = pd.ExcelFile(io.BytesIO(content))
        sheet_name = _select_pandas_sheet(xls)
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Eski Excel (.xls) dosyası okunamadı. Gerekirse .xlsx olarak yükleyin. Hata: {exc}",
        ) from exc

    rows = [
        [_normalize_cell(value) for value in row]
        for row in df.where(pd.notna(df), None).values.tolist()
    ]
    return str(sheet_name), _trim_empty_edges(rows)


def _select_worksheet(wb: openpyxl.Workbook):
    for ws in wb.worksheets:
        if ws.title.strip().upper() == "DATA":
            return ws
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 12), values_only=True):
            values = [str(v).upper() for v in row if v is not None]
            if any("ANKET ADI" in value for value in values):
                return ws
    return wb.worksheets[0]


def _select_pandas_sheet(xls: pd.ExcelFile) -> str:
    for name in xls.sheet_names:
        if str(name).strip().upper() == "DATA":
            return name
    return xls.sheet_names[0]


def _decode_csv_text(content: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "cp1254", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _trim_empty_edges(rows: list[list[Any]]) -> list[list[Any]]:
    trimmed = [row for row in rows if any(_is_present(value) for value in row)]
    if not trimmed:
        return []
    max_len = max(len(row) for row in trimmed)
    return [row + [""] * (max_len - len(row)) for row in trimmed]


def _detect_title(rows: list[list[Any]]) -> str | None:
    for row in rows[:8]:
        for value in row:
            text = str(value).strip()
            if "ANKET ADI" in text.upper():
                return re.sub(r"^\s*ANKET ADI\s*:?\s*", "", text, flags=re.IGNORECASE).strip()
    return None


def _detect_header_row(rows: list[list[Any]]) -> int | None:
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:25]):
        values = [str(value).strip() for value in row if _is_present(value)]
        if not values:
            continue

        lowered = " ".join(values).lower()
        question_count = sum(1 for value in values if _question_no(value)[0] is not None)
        score = 0
        if values[0].lower().startswith("no"):
            score += 5
        if "tarih" in lowered:
            score += 3
        if question_count:
            score += question_count * 2
        if len(values) >= 5:
            score += 2
        if index > 0 and _count_code_like_cells(rows[index - 1]) >= 2:
            score += 3
        if "kodlarken" in lowered:
            score -= 8
        if "anket adi" in lowered:
            score -= 8

        if score > best_score:
            best_score = score
            best_index = index

    return best_index if best_score >= 5 else None


def _count_code_like_cells(row: list[Any]) -> int:
    return sum(1 for value in row if _parse_code_map(str(value)).get("options") or str(value).strip() in {"Metin", "Sayı", "Tarih"})


def _dedupe_headers(row: list[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        name = str(value).strip() if _is_present(value) else f"Kolon {index}"
        name = re.sub(r"\s+", " ", name)
        if not name:
            name = f"Kolon {index}"
        base = name
        counts[base] = counts.get(base, 0) + 1
        if counts[base] > 1:
            name = f"{base} ({counts[base]})"
        headers.append(name)
    return headers


def _build_data_rows(raw_rows: list[list[Any]], headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in raw_rows:
        padded = raw + [""] * (len(headers) - len(raw))
        values = padded[: len(headers)]
        if not any(_is_present(value) for value in values):
            continue
        rows.append({header: _jsonable(value) for header, value in zip(headers, values)})
    return rows


def _build_columns(headers: list[str], code_row: list[Any], data_rows: list[dict[str, Any]]) -> list[ParsedColumn]:
    columns: list[ParsedColumn] = []
    for index, header in enumerate(headers):
        values = [row.get(header) for row in data_rows]
        present_values = [value for value in values if _is_present(value)]
        code_hint = str(code_row[index]).strip() if index < len(code_row) and _is_present(code_row[index]) else ""
        code_data = _parse_code_map(code_hint)
        dtype = _infer_dtype(present_values, code_hint)
        semantic_type = _semantic_type(header, dtype, code_data["options"])
        unique_values = {_stable_key(value) for value in present_values}
        missing_count = len(values) - len(present_values)
        sample_values = _sample_values(present_values)
        columns.append(
            ParsedColumn(
                name=header,
                original_name=header,
                dtype=dtype,
                semantic_type=semantic_type,
                missing_count=missing_count,
                missing_pct=round((missing_count / len(values)) * 100, 2) if values else 0.0,
                unique_count=len(unique_values),
                sample_values=sample_values,
                code_map=code_data["options"],
                order_index=index,
            )
        )
    return columns


def _build_questions(columns: list[ParsedColumn]) -> list[ParsedQuestion]:
    questions: list[ParsedQuestion] = []
    for column in columns:
        question_no, question_text = _question_no(column.name)
        is_demographic = column.semantic_type == "demographic"
        is_likert, scale_type = _detect_likert(column.code_map)
        is_open_text = column.dtype == "text" and not column.code_map and not is_demographic

        if question_no:
            question_type = "likert" if is_likert else "single_choice" if column.code_map else "open_text" if is_open_text else column.dtype
        elif is_demographic:
            question_type = "demographic"
        elif column.code_map:
            question_type = "coded"
        else:
            question_type = column.dtype

        questions.append(
            ParsedQuestion(
                column_name=column.name,
                question_no=question_no,
                question_text=question_text,
                question_type=question_type,
                scale_type=scale_type,
                options=column.code_map,
                is_likert=is_likert,
                is_demographic=is_demographic,
                is_open_text=is_open_text,
                order_index=column.order_index,
            )
        )
    return questions


def _build_report(
    title: str,
    columns: list[ParsedColumn],
    questions: list[ParsedQuestion],
    rows: list[dict[str, Any]],
) -> SurveyReportData:
    question_by_column = {question.column_name: question for question in questions}
    quality_issues = _quality_issues(columns, question_by_column, rows)
    question_metrics = []
    likert_values: list[float] = []

    for column in columns:
        question = question_by_column[column.name]
        values = [row.get(column.name) for row in rows]
        present_values = [value for value in values if _is_present(value)]
        distribution = _distribution(present_values, column.code_map)
        numeric_values = [_to_float(value) for value in present_values]
        numeric_values = [value for value in numeric_values if value is not None]
        average = round(sum(numeric_values) / len(numeric_values), 4) if numeric_values else None
        satisfaction_score = None
        if question.is_likert and numeric_values:
            satisfaction_score = round((sum(numeric_values) / len(numeric_values)) * 20, 2)
            likert_values.extend(numeric_values)

        question_metrics.append(
            {
                "column_name": column.name,
                "question_no": question.question_no,
                "question_text": question.question_text,
                "question_type": question.question_type,
                "response_count": len(present_values),
                "missing_count": column.missing_count,
                "missing_pct": column.missing_pct,
                "average": average,
                "satisfaction_score": satisfaction_score,
                "distribution": distribution,
            }
        )

    overall_satisfaction = None
    if likert_values:
        overall_satisfaction = round((sum(likert_values) / len(likert_values)) * 20, 2)

    summary = {
        "title": title,
        "response_count": len(rows),
        "column_count": len(columns),
        "question_count": len(questions),
        "likert_question_count": sum(1 for question in questions if question.is_likert),
        "open_text_question_count": sum(1 for question in questions if question.is_open_text),
        "overall_satisfaction": overall_satisfaction,
    }
    metrics = {
        "question_metrics": question_metrics,
        "top_missing_columns": sorted(
            [
                {"column_name": column.name, "missing_pct": column.missing_pct}
                for column in columns
                if column.missing_count > 0
            ],
            key=lambda item: item["missing_pct"],
            reverse=True,
        )[:10],
    }
    report_text = _report_text(summary, quality_issues)
    return SurveyReportData(summary=summary, metrics=metrics, quality_issues=quality_issues, report_text=report_text)


def _parse_code_map(text: str) -> dict[str, Any]:
    options: dict[str, str] = {}
    if not text:
        return {"options": options}
    normalized = str(text).replace("\\n", "\n")
    pattern = re.compile(
        r"(?P<code>[\"']{0,2}[0-9A-Za-zÇĞİÖŞÜçğıöşü]+[\"']{0,2})\s*:+\s*",
        flags=re.UNICODE,
    )
    matches = list(pattern.finditer(normalized))
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        code = re.sub(r"[^0-9A-Za-zÇĞİÖŞÜçğıöşü]", "", match.group("code")).strip()
        label = normalized[start:end].strip().strip("\"' |;\n\r\t")
        if code and label:
            options[code] = label
    return {"options": options}


def _infer_dtype(values: list[Any], code_hint: str) -> str:
    hint = code_hint.strip().lower()
    if hint == "tarih":
        return "datetime"
    if hint in {"metin", "text"}:
        return "text"
    if hint in {"sayı", "sayi", "number"}:
        return "numeric"
    if not values:
        return "unknown"
    if all(isinstance(value, datetime) for value in values):
        return "datetime"
    numeric_count = sum(1 for value in values if _to_float(value) is not None)
    if numeric_count / len(values) >= 0.9:
        return "numeric"
    datetime_count = sum(1 for value in values if _looks_like_datetime(value))
    if datetime_count / len(values) >= 0.9:
        return "datetime"
    if len({_stable_key(value) for value in values}) <= max(20, len(values) * 0.2):
        return "categorical"
    return "text"


def _semantic_type(header: str, dtype: str, code_map: dict[str, str]) -> str:
    text = _ascii_fold(header).lower()
    if any(keyword in text for keyword in DEMOGRAPHIC_KEYWORDS):
        return "demographic"
    if _question_no(header)[0] is not None:
        return "question"
    if code_map:
        return "coded"
    return dtype


def _detect_likert(options: dict[str, str]) -> tuple[bool, str | None]:
    if not options:
        return False, None
    labels = " ".join(options.values()).lower()
    codes = set(options.keys())
    if {"1", "2", "3", "4", "5"}.issubset(codes) and "memnun" in labels:
        return True, "satisfaction_5"
    if {"1", "2", "3"}.issubset(codes) and ("memnun" in labels or "kararsız" in labels or "kararsiz" in labels):
        return True, "satisfaction_3"
    return False, None


def _question_no(header: str) -> tuple[str | None, str]:
    text = str(header).strip()
    match = re.match(r"^(\d+)[\.)]\s*(.+)$", text, flags=re.DOTALL)
    if not match:
        return None, text
    return match.group(1), match.group(2).strip()


def _distribution(values: list[Any], code_map: dict[str, str]) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    labels: dict[str, str] = {}
    for value in values:
        key = _option_key(value)
        counts[key] = counts.get(key, 0) + 1
        labels[key] = code_map.get(key, str(value))
    total = sum(counts.values())
    return [
        {"value": key, "label": labels[key], "count": count, "pct": round((count / total) * 100, 2) if total else 0.0}
        for key, count in sorted(counts.items(), key=lambda item: item[1], reverse=True)
    ]


def _quality_issues(
    columns: list[ParsedColumn],
    question_by_column: dict[str, ParsedQuestion],
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    duplicate_base_names = _duplicate_base_names([column.original_name for column in columns])
    for name, count in duplicate_base_names.items():
        issues.append({"type": "duplicate_column", "severity": "warning", "message": f"'{name}' kolonu {count} kez tekrar ediyor"})

    for column in columns:
        if column.missing_pct >= 30:
            issues.append(
                {
                    "type": "high_missing_rate",
                    "severity": "warning",
                    "column": column.name,
                    "message": f"{column.name} kolonunda eksik veri oranı %{column.missing_pct}",
                }
            )
        if column.code_map:
            unexpected = sorted(
                {
                    _option_key(row.get(column.name))
                    for row in rows
                    if _is_present(row.get(column.name)) and _option_key(row.get(column.name)) not in column.code_map
                }
            )
            if unexpected:
                issues.append(
                    {
                        "type": "unexpected_code",
                        "severity": "warning",
                        "column": column.name,
                        "values": unexpected[:20],
                        "message": f"{column.name} kolonunda kodlama sözlüğünde olmayan değerler var",
                    }
                )

    return issues


def _report_text(summary: dict[str, Any], quality_issues: list[dict[str, Any]]) -> str:
    satisfaction = summary.get("overall_satisfaction")
    satisfaction_text = f"%{satisfaction}" if satisfaction is not None else "hesaplanamadı"
    return (
        f"{summary['title']} için {summary['response_count']} yanıt ve {summary['question_count']} alan işlendi. "
        f"Likert/memnuniyet sorusu sayısı {summary['likert_question_count']}. "
        f"Genel memnuniyet skoru {satisfaction_text}. "
        f"Veri kalite uyarısı sayısı {len(quality_issues)}."
    )


def _detect_department(title: str) -> str | None:
    marker = " MÜDÜRLÜĞÜ"
    if marker in title:
        return title.split(marker)[0].strip() + marker
    return None


def _detect_period(filename: str) -> str | None:
    lower = _ascii_fold(filename).lower()
    if "ceyrek" in lower:
        return "quarter"
    if "yariyil" in lower or "yarıyıl" in lower or "6 aylik" in lower:
        return "half_year"
    if "yillik" in lower or "yıllık" in lower:
        return "year"
    return None


def _detect_quarter(filename: str) -> str | None:
    lower = _ascii_fold(filename).lower()
    for value in ("1", "2", "3", "4"):
        if f"{value}. ceyrek" in lower or f"{value}.ceyrek" in lower or f"{value} ceyrek" in lower:
            return value
    return None


def _detect_year(filename: str) -> int | None:
    match = re.search(r"(20\d{2})", filename)
    return int(match.group(1)) if match else None


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _jsonable(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def _is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return str(value).strip() != ""


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _looks_like_datetime(value: Any) -> bool:
    if isinstance(value, datetime):
        return True
    text = str(value).strip()
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}", text) or re.match(r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", text))


def _option_key(value: Any) -> str:
    number = _to_float(value)
    if number is not None and number.is_integer():
        return str(int(number))
    return str(value).strip()


def _stable_key(value: Any) -> str:
    return _option_key(value).lower()


def _sample_values(values: list[Any], limit: int = 5) -> list[Any]:
    samples: list[Any] = []
    seen: set[str] = set()
    for value in values:
        key = _stable_key(value)
        if key in seen:
            continue
        samples.append(_jsonable(value))
        seen.add(key)
        if len(samples) >= limit:
            break
    return samples


def _duplicate_base_names(names: list[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for name in names:
        counts[name] = counts.get(name, 0) + 1
    return {name: count for name, count in counts.items() if count > 1}


def _ascii_fold(value: str) -> str:
    return (
        str(value)
        .replace("Ç", "C")
        .replace("ç", "c")
        .replace("Ğ", "G")
        .replace("ğ", "g")
        .replace("İ", "I")
        .replace("ı", "i")
        .replace("Ö", "O")
        .replace("ö", "o")
        .replace("Ş", "S")
        .replace("ş", "s")
        .replace("Ü", "U")
        .replace("ü", "u")
    )
